import datetime
import openpyxl
import re
from openpyxl.utils import get_column_letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.shortcuts import render, redirect
from urllib.parse import quote

from .forms import RegistroForm
from .models import Comida
from django.db.models import Q, F
from django.http import JsonResponse, HttpResponse
from .models import Registro, Casino
# Create your views here.

def login_view(request):
    # Si el usuario ya está autenticado, redirigirlo directamente
    print("Usuario en sesión:", request.user)
    if request.user.is_authenticated:
        if request.user.es_administrador:
            return redirect("registro_app:reporte_mensual")
        else:
            return redirect("registro_app:registro")

    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")

        user = authenticate(request, username=username, password=password)

        print("Intentando login con:", username, password)
        print("Resultado authenticate:", user)

        if user is not None:
            login(request, user)
            if user.es_administrador:
                return redirect("registro_app:reporte_mensual")
            else:
                return redirect("registro_app:registro")
        else:
            messages.error(request, "Usuario o contraseña incorrectos.")

    return render(request, "login.html")

def logout_view(request):
    logout(request)  # elimina la sesión del usuario
    messages.success(request, "Has cerrado sesión correctamente.")
    return redirect("login")  # cambia "login" por el nombre de tu url de login

@login_required(login_url='login')  # redirige al login si no está autenticado
def registrar(request):
    usuario = request.user
    comidas = Comida.objects.all()  
    # if not usuario.is_authenticated:
    #     return redirect('login')
    form = RegistroForm(request.POST)
    return render(request,"registrar.html",{'form': form, 'usuario': usuario, 'comidas': comidas})

def crear_registro(request):
    fecha_hoy = datetime.date.today()
    if request.method == 'POST':
        form = RegistroForm(request.POST)
        if form.is_valid():
            documento = form.cleaned_data['documento']
            comidas_ids = request.POST.getlist('comida')
            print("Comidas seleccionadas:", comidas_ids)
            # Creamos un objeto base sin guardar
            registro_base = form.save(commit=False)
            
            # Validación: ya existe registro para este documento en estas comidas hoy?
            registros_existentes = Registro.objects.filter(
                documento=documento,
                fecha_hora__date=fecha_hoy,
                comida_id__in=comidas_ids,
            )

            if registros_existentes.exists():
                # Ya está registrado para al menos una de esas comidas hoy
                messages.error(request, "Ya te has anotado hoy en una de las comidas seleccionadas.")
                return redirect('registro_app:registro')
            else:
                for comida_id in comidas_ids:
                    comida = Comida.objects.get(id=comida_id)
                    registro = Registro(
                        **{
                            campo.name: getattr(registro_base, campo.name)
                            for campo in Registro._meta.fields
                            if campo.name not in ('id', 'comida', 'fecha_hora')
                        }
                    )
                    registro.comida = comida
                    registro.save()
                return redirect('registro_app:registro_exitoso')
        else:
            print("Formulario no válido:", form.errors)
    else:
        form = RegistroForm()
    
    return render(request, 'registro/crear_registro.html', {'form': form})

@login_required(login_url='login')  # redirige al login si no está autenticado
def registro_exitoso(request):
    logout(request)  # elimina la sesión del usuario
    return render(request, 'registro_exitoso.html')

@login_required(login_url='login')  # redirige al login si no está autenticado    
def reporte_mensual_view(request):
    usuario=request.user
    if usuario.es_administrador == False:
        return redirect('registro_app:registro')
    context = {
        'comidas': Comida.objects.all(),
        'casinos': Casino.objects.all(),
        'fecha_hoy': datetime.date.today().strftime('%d/%m/%Y'),
        'hero': True,
    }
    
    return render(request,"reporte_mensual.html",{'data': context})

def config_comidas(request):
    # Habilita o deshbilita comidas disponibles para el registro segun su campo 'habilitado'
    if request.method == 'POST':
        comidas_ids = request.POST.getlist('comida')
        print("Comidas seleccionadas para habilitar:", comidas_ids)
        # Primero deshabilitamos todas
        Comida.objects.update(habilitado=False)
        # Luego habilitamos las seleccionadas
        Comida.objects.filter(id__in=comidas_ids).update(habilitado=True)
        messages.success(request, "Configuración de comidas actualizada.")
        return redirect('registro_app:reporte_mensual')

def registro_datatable(request):
    draw = request.GET.get('draw')
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', 10))
    order_column_index = int(request.GET.get('order[0][column]', 0))
    order_direction = request.GET.get('order[0][dir]', 'asc')
    search_value = request.GET.get('search[value]', None)

    filtro_comida = request.GET.get('comida')
    filtro_casino = request.GET.get('casino')
    filtro_fecha  = request.GET.get('fecha')

    # Mapeo de índices a columnas del datatable
    column_mapping = {
        0: 'fecha_hora', 
        1: 'apellido', 
        2: 'nombre', 
        3: 'documento', 
        4: 'casino', 
        5: 'comida',
    }

    order_column = column_mapping.get(order_column_index, 'apellido')
    if order_direction == 'desc':
        order_column = F(order_column).asc(nulls_last=True)
    else:
        order_column = F(order_column).desc(nulls_last=True)

    # Construimos dinámicamente las condiciones de búsqueda utilizando Q objects
    conditions = Q()
    if search_value:
        fields = [
            'documento', 
        ]
        search_terms = search_value.split()
        for term in search_terms:
            term_conditions = Q()
            for field in fields:
                term_conditions |= Q(**{f"{field}__icontains": term})

            conditions &= term_conditions

    filtered_data = Registro.objects.filter(conditions)
    
    # Aplicamos filtros dinámicos
    if filtro_comida:
        filtered_data = filtered_data.filter(comida=filtro_comida)
    if filtro_casino:
        filtered_data = filtered_data.filter(casino=filtro_casino)
    if filtro_fecha:
        try:
            fecha_obj = datetime.datetime.strptime(filtro_fecha, '%Y-%m-%d').date()
            filtered_data = filtered_data.filter(fecha_hora__date=fecha_obj)
        except ValueError:
            print("Formato de fecha inválido:", filtro_fecha)
    else:
        hoy = datetime.date.today()
        filtered_data = filtered_data.filter(fecha_hora__date=hoy)
        
    filtered_data = filtered_data.order_by(order_column)
    total_records = filtered_data.count()

    # Paginación
    data = [
        item.to_json()
        for item in filtered_data[start: start + length]
    ]   

    return JsonResponse({
        'draw': draw,
        'recordsTotal': total_records,
        'recordsFiltered': total_records,
        'data': data
    })

def export_excel(request):
    comida = request.GET.get("comida")
    casino = request.GET.get("casino")
    fecha = request.GET.get("fecha")

    registros = Registro.objects.all()

    if comida:
        registros = registros.filter(comida=comida)
    if casino:
        registros = registros.filter(casino=casino)
    if fecha:
        registros = registros.filter(fecha_hora__date=fecha)

    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Registros"

    # Encabezados
    headers = ["Fecha", "Apellido", "Nombre", "Documento", "Casino", "Comida"]
    ws.append(headers)

    # Filas
    for r in registros:
        ws.append([
            r.fecha_hora.strftime("%d/%m/%Y %H:%M"),
            r.apellido,
            r.nombre,
            r.documento or "Sin DNI",
            r.casino.nombre if r.casino else "",
            r.comida.nombre if r.comida else "",
        ])

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    # Respuesta HTTP
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    disposition, _ = build_filename(comida, casino, fecha, ext="xlsx")
    response["Content-Disposition"] = disposition
    wb.save(response)
    return response

def export_pdf(request):
    comida = request.GET.get("comida")
    casino = request.GET.get("casino")
    fecha = request.GET.get("fecha")

    registros = Registro.objects.all()

    if comida:
        registros = registros.filter(comida=comida)
    if casino:
        registros = registros.filter(casino=casino)
    if fecha:
        registros = registros.filter(fecha_hora__date=fecha)

    # Crear PDF
    response = HttpResponse(content_type="application/pdf")
    disposition, _ = build_filename(comida, casino, fecha, ext="pdf")
    response["Content-Disposition"] = disposition

    doc = SimpleDocTemplate(response, pagesize=A4)
    styles = getSampleStyleSheet()
    elementos = []

    # Encabezado
    elementos.append(Paragraph("Listado de Registros", styles["Title"]))

    # Tabla
    data = [["Fecha", "Apellido", "Nombre", "Documento", "Casino", "Comida"]]
    for r in registros:
        data.append([
            r.fecha_hora.strftime("%d/%m/%Y %H:%M"),
            r.apellido,
            r.nombre,
            r.documento or "Sin DNI",
            r.casino,
            r.comida,
        ])

    table = Table(data)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.grey),
        ("TEXTCOLOR", (0,0), (-1,0), colors.whitesmoke),
        ("ALIGN", (0,0), (-1,-1), "CENTER"),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("BOTTOMPADDING", (0,0), (-1,0), 10),
        ("GRID", (0,0), (-1,-1), 1, colors.black),
    ]))
    elementos.append(table)

    doc.build(elementos)
    return response

def sanitize_filename(text):
    """Quita caracteres no válidos para nombres de archivo"""
    if not text:
        return ""
    # reemplaza espacios por guiones bajos y elimina caracteres raros
    return re.sub(r'[^A-Za-z0-9_\-]', '', text.replace(" ", "_"))

def build_filename(comida=None, casino=None, fecha=None, default="Registro", ext="pdf"):
    parts = []

    if comida:
        comida_text = Comida.objects.get(id=comida).nombre
        parts.append(sanitize_filename(comida_text))
    if casino:
        casino_text = Casino.objects.get(id=casino).nombre
        parts.append(sanitize_filename(casino_text))
    if not parts:
        parts.append(default)

    if fecha:
        try:
            # Intentar parsear si viene como YYYY-MM-DD
            dt = datetime.strptime(fecha, "%Y-%m-%d")
            date_str = dt.strftime("%Y-%m-%d")
        except Exception:
            # Si no se puede, usar la cadena tal cual sanitizada
            date_str = sanitize_filename(fecha)
    else:
        date_str = datetime.now().strftime("%Y-%m-%d")

    parts.append(date_str)
    filename = "-".join(parts) + "." + ext

    # Cabecera compatible con UTF-8
    content_disposition = f'attachment; filename="{filename}"; filename*=UTF-8\'\'{quote(filename)}'
    return content_disposition, filename