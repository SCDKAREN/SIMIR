import openpyxl
import re
import datetime
from django.utils import timezone
from django.core.cache import cache
from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect
from django.contrib import messages

from openpyxl.utils import get_column_letter
from reportlab.lib.styles import getSampleStyleSheet
from urllib.parse import quote
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4

from .models import Registro, Comida, Casino
from .forms import RegistroForm

# region REGISTRO
def crear_registro(request):
    fecha_hoy = datetime.date.today()
    if request.method == 'POST':
        form = RegistroForm(request.POST)
        if form.is_valid():
            documento = form.cleaned_data['documento']
            comidas_ids = request.POST.getlist('comida')
            print("Comidas seleccionadas:", comidas_ids)
            # Creamos un objeto base sin guardar
            registro_base = form.save(commit=False)
            
            # Validación: ya existe registro para este documento en estas comidas hoy?
            registros_existentes = Registro.objects.filter(
                documento=documento,
                fecha_hora__date=fecha_hoy,
                comida_id__in=comidas_ids,
            )

            if registros_existentes.exists():
                # Ya está registrado para al menos una de esas comidas hoy
                messages.error(request, "Ya te has anotado hoy en una o más de las comidas seleccionadas.")
                return redirect('registro_app:registro')
            else:
                for comida_id in comidas_ids:
                    comida = Comida.objects.get(id=comida_id)
                    registro = Registro(
                        **{
                            campo.name: getattr(registro_base, campo.name)
                            for campo in Registro._meta.fields
                            if campo.name not in ('id', 'comida', 'fecha_hora')
                        }
                    )
                    registro.comida = comida
                    registro.save()
                return redirect('registro_app:registro_exitoso')
        else:
            print("Formulario no válido:", form.errors)
    else:
        form = RegistroForm()
    # Redirect por defecto
    return redirect('registro_app:registro')

def config_comidas(request):
    # Habilita o deshbilita comidas disponibles para el registro segun su campo 'habilitado'
    if request.method == 'POST':
        comidas_ids = request.POST.getlist('comida')
        # Primero deshabilitamos todas
        Comida.objects.update(habilitado=False)
        # Luego habilitamos las seleccionadas
        Comida.objects.filter(id__in=comidas_ids).update(habilitado=True)
        return redirect('registro_app:reporte_mensual')
    

def activar_formulario(request):
    if not request.user.is_authenticated or not request.user.es_administrador:
        return JsonResponse({'success': False, 'message': 'No autorizado'}, status=403)

    expiracion = datetime.datetime.now() + datetime.timedelta(minutes=30)
    
    # Guardamos en el cache global (clave accesible por todas las sesiones)
    cache.set('formulario_activado_hasta', expiracion, timeout=30*60)

    return JsonResponse({
        'success': True,
        'expira': expiracion.strftime("%H:%M"),
        'message': 'Formulario habilitado globalmente por 30 minutos'
    })

# region EXPORTAR EXCEL Y PDF
def export_excel(request):
    comida = request.GET.get("comida")
    casino = request.GET.get("casino")
    fecha = request.GET.get("fecha")

    registros = Registro.objects.all()

    if comida:
        registros = registros.filter(comida=comida)
    if casino:
        registros = registros.filter(casino=casino)
    if fecha:
        registros = registros.filter(fecha_hora__date=fecha)

    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Registros"

    # Encabezados
    headers = ["Fecha", "Apellido", "Nombre", "Documento", "Casino", "Comida"]
    ws.append(headers)

    # Filas
    for r in registros:
        ws.append([
            r.fecha_hora.strftime("%d/%m/%Y %H:%M"),
            r.apellido,
            r.nombre,
            r.documento or "Sin DNI",
            r.casino.nombre if r.casino else "",
            r.comida.nombre if r.comida else "",
        ])

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    # Respuesta HTTP
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    disposition, _ = build_filename(comida, casino, fecha, ext="xlsx")
    response["Content-Disposition"] = disposition
    wb.save(response)
    return response

def export_pdf(request):
    comida = request.GET.get("comida")
    casino = request.GET.get("casino")
    fecha = request.GET.get("fecha")

    registros = Registro.objects.all()

    if comida:
        registros = registros.filter(comida=comida)
    if casino:
        registros = registros.filter(casino=casino)
    if fecha:
        registros = registros.filter(fecha_hora__date=fecha)

    # Crear PDF
    response = HttpResponse(content_type="application/pdf")
    disposition, _ = build_filename(comida, casino, fecha, ext="pdf")
    response["Content-Disposition"] = disposition

    doc = SimpleDocTemplate(response, pagesize=A4)
    styles = getSampleStyleSheet()
    elementos = []

    # Encabezado
    elementos.append(Paragraph("Listado de Registros", styles["Title"]))

    # Tabla
    data = [["Fecha", "Apellido", "Nombre", "Documento", "Casino", "Comida"]]
    for r in registros:
        data.append([
            r.fecha_hora.strftime("%d/%m/%Y %H:%M"),
            r.apellido,
            r.nombre,
            r.documento or "Sin DNI",
            r.casino,
            r.comida,
        ])

    table = Table(data)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.grey),
        ("TEXTCOLOR", (0,0), (-1,0), colors.whitesmoke),
        ("ALIGN", (0,0), (-1,-1), "CENTER"),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("BOTTOMPADDING", (0,0), (-1,0), 10),
        ("GRID", (0,0), (-1,-1), 1, colors.black),
    ]))
    elementos.append(table)

    doc.build(elementos)
    return response

def sanitize_filename(text):
    """Quita caracteres no válidos para nombres de archivo"""
    if not text:
        return ""
    # reemplaza espacios por guiones bajos y elimina caracteres raros
    return re.sub(r'[^A-Za-z0-9_\-]', '', text.replace(" ", "_"))

def build_filename(comida=None, casino=None, fecha=None, default="Registro", ext="pdf"):
    parts = []

    if comida:
        comida_text = Comida.objects.get(id=comida).nombre
        parts.append(sanitize_filename(comida_text))
    if casino:
        casino_text = Casino.objects.get(id=casino).nombre
        parts.append(sanitize_filename(casino_text))
    if not parts:
        parts.append(default)

    if fecha:
        try:
            # Intentar parsear si viene como YYYY-MM-DD
            dt = datetime.strptime(fecha, "%Y-%m-%d")
            date_str = dt.strftime("%Y-%m-%d")
        except Exception:
            # Si no se puede, usar la cadena tal cual sanitizada
            date_str = sanitize_filename(fecha)
    else:
        date_str = datetime.now().strftime("%Y-%m-%d")

    parts.append(date_str)
    filename = "-".join(parts) + "." + ext

    # Cabecera compatible con UTF-8
    content_disposition = f'attachment; filename="{filename}"; filename*=UTF-8\'\'{quote(filename)}'
    return content_disposition, filename

